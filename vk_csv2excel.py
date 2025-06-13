#!/usr/bin/env python3
"""
Multiple CSV/TSV to Excel Importer
This script creates an Excel workbook with separate sheets for each CSV/TSV file.

Usage:
    python script.py --input file1.csv file2.tsv [file3.csv ...] --output result.xlsx
    python script.py -i file1.csv file2.tsv -o result.xlsx
    python script.py -i directory_with_csv_files -o result.xlsx

Requirements:
- pandas
- openpyxl

Install with: pip install pandas openpyxl
"""

import os
import sys
import pandas as pd
import glob
import argparse
from pathlib import Path
import re
import numpy as np
from openpyxl.styles import PatternFill, Color, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


def detect_delimiter(file_path):
    """
    Detect if a file is CSV or TSV based on extension and content analysis.
    Returns the appropriate delimiter.
    """
    # Check if file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
        
    # Check extension first
    if file_path.lower().endswith('.csv'):
        return ','
    elif file_path.lower().endswith('.tsv'):
        return '\t'
    
    # If extension doesn't clearly indicate, analyze content
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            sample = f.readline()
    except Exception as e:
        print(f"Warning: Could not read first line of {file_path} for delimiter detection: {e}")
        # Default to comma if we can't read the file
        return ','
    
    # Count potential delimiters
    tab_count = sample.count('\t')
    comma_count = sample.count(',')
    semicolon_count = sample.count(';')  # Common in European CSV files
    pipe_count = sample.count('|')       # Sometimes used in data exports
    
    # Determine the most likely delimiter
    counts = {
        '\t': tab_count,
        ',': comma_count,
        ';': semicolon_count,
        '|': pipe_count
    }
    
    # Return the delimiter with highest count
    max_delimiter = max(counts.items(), key=lambda x: x[1])
    
    if max_delimiter[1] == 0:
        # If no clear delimiter is found, default to comma
        print(f"Warning: No clear delimiter found in {file_path}, defaulting to comma")
        return ','
    
    print(f"Detected delimiter '{max_delimiter[0]}' for file {os.path.basename(file_path)}")
    return max_delimiter[0]


def is_numeric_column(series):
    """Check if a pandas Series contains numeric data."""
    # Try to convert to numeric, if most values convert successfully, it's numeric
    numeric_count = pd.to_numeric(series, errors='coerce').notna().sum()
    return numeric_count > len(series) * 0.7  # If >70% values are numeric


def apply_heatmap(worksheet, df, skip_columns=None):
    """Apply heatmap coloring to numeric columns."""
    if skip_columns is None:
        skip_columns = []
    
    # Process each column
    for i, col_name in enumerate(df.columns):
        # Skip columns in the skip list
        if col_name in skip_columns:
            continue
            
        # Get column letter
        col_letter = get_column_letter(i+1)
        
        # Check if column contains numeric data
        if is_numeric_column(df[col_name]):
            # Set color scale (red-yellow-green)
            color_scale = ColorScaleRule(
                start_type='min', start_color='F8696B',  # red
                mid_type='percentile', mid_value=50, mid_color='FFEB84',  # yellow
                end_type='max', end_color='63BE7B'  # green
            )
            
            # Apply color scale to data range (excluding header)
            data_range = f"{col_letter}2:{col_letter}{len(df)+1}"
            worksheet.conditional_formatting.add(data_range, color_scale)
            
            print(f"  Applied heatmap to column: {col_name}")


def import_files_to_excel(file_paths, output_excel=None, horizontal_headers=False, apply_heatmap_coloring=False, skip_heatmap_cols=None):
    """
    Import multiple CSV/TSV files into a single Excel file with separate sheets.
    
    Args:
        file_paths: List of file paths to import
        output_excel: Path to save the Excel file (optional)
        horizontal_headers: If True, headers will remain horizontal (default is vertical)
        apply_heatmap_coloring: If True, numeric columns will get heatmap colors
        skip_heatmap_cols: List of column names to exclude from heatmap coloring
    
    Returns:
        Tuple of (success_count, total_count, output_path)
    """
    if not file_paths:
        print("No files to process.")
        return 0, 0, None
    
    # Create a new Excel writer
    if not output_excel:
        # Use the directory of the first file if no output path is specified
        first_dir = os.path.dirname(file_paths[0])
        output_excel = os.path.join(first_dir, "combined_data.xlsx")
    
    # Ensure the output directory exists
    output_dir = os.path.dirname(os.path.abspath(output_excel))
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
            print(f"Created directory: {output_dir}")
        except Exception as e:
            print(f"Error creating directory {output_dir}: {e}")
            return 0, len(file_paths), None
    
    # Create the Excel writer with error handling
    try:
        writer = pd.ExcelWriter(output_excel, engine='openpyxl')
    except Exception as e:
        print(f"Error creating Excel file at {output_excel}: {e}")
        print(f"Please check if you have write permissions to this location or if the file is open in another program.")
        return 0, len(file_paths), None
    
    imported_count = 0
    skipped_files = []
    
    # Process each file
    for file_path in file_paths:
        try:
            print(f"Processing: {file_path}")
            
            # Get sheet name from filename (without extension)
            sheet_name = Path(file_path).stem
            
            # Excel has a 31 character limit for sheet names
            if len(sheet_name) > 31:
                old_name = sheet_name
                sheet_name = sheet_name[:31]
                print(f"  Note: Sheet name truncated from '{old_name}' to '{sheet_name}'")
            
            # Replace invalid characters in sheet name
            invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
            for char in invalid_chars:
                if char in sheet_name:
                    sheet_name = sheet_name.replace(char, '_')
                    print(f"  Note: Invalid character '{char}' in sheet name replaced with '_'")
            
            # Determine delimiter
            delimiter = detect_delimiter(file_path)
            
            # Read the file
            try:
                df = pd.read_csv(file_path, delimiter=delimiter, encoding='utf-8', 
                                on_bad_lines='warn', low_memory=False)
            except UnicodeDecodeError:
                # Try with different encodings if UTF-8 fails
                try:
                    df = pd.read_csv(file_path, delimiter=delimiter, encoding='latin1', 
                                    on_bad_lines='warn', low_memory=False)
                    print(f"  Note: File {os.path.basename(file_path)} used Latin-1 encoding")
                except Exception as e2:
                    raise Exception(f"Encoding error: {e2}. Try specifying the correct encoding.")
            except Exception as e:
                raise Exception(f"Failed to read file: {e}")
            
            # Check if DataFrame is empty
            if df.empty:
                print(f"  Warning: File {os.path.basename(file_path)} is empty or has no valid data")
                
            # Write to Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get worksheet
            worksheet = writer.sheets[sheet_name]
            
            if not horizontal_headers:
                # Set headers vertical (rotated 90 degrees)
                for i, col in enumerate(df.columns):
                    cell = worksheet.cell(row=1, column=i+1)
                    # Create new Alignment object with correct attributes
                    cell.alignment = Alignment(textRotation=90, vertical='center', horizontal='center')
                    
                    # Find the maximum length in the column data (excluding headers)
                    if len(df) > 0:
                        max_data_len = df[col].astype(str).map(len).max() + 2  # add space
                    else:
                        max_data_len = 10  # default width
                    
                    # Set width based on data (not header, since header is now vertical)
                    worksheet.column_dimensions[get_column_letter(i+1)].width = min(max_data_len, 50)  # cap at 50
                
                # Increase the height of the header row to accommodate vertical text
                worksheet.row_dimensions[1].height = 150
            else:
                # Keep headers horizontal (default Excel behavior)
                for i, col in enumerate(df.columns):
                    # Find the maximum length in the column including header
                    if len(df) > 0:
                        # Handle potential NaN values by converting to string first
                        max_data_len = max(len(str(x)) for x in df[col].fillna('').astype(str))
                    else:
                        max_data_len = 0
                        
                    max_len = max(
                        max_data_len,  # max length of column data
                        len(str(col))  # length of column name
                    ) + 2  # add a little extra space
                    
                    # Excel's column width is different from string length
                    worksheet.column_dimensions[get_column_letter(i+1)].width = min(max_len, 50)  # cap at 50
            
            # Apply heatmap if requested
            if apply_heatmap_coloring:
                print(f"Applying heatmap to numeric columns in {sheet_name}...")
                apply_heatmap(worksheet, df, skip_heatmap_cols)
            
            imported_count += 1
            print(f"Successfully imported: {file_path} → Sheet: {sheet_name}")
            
        except Exception as e:
            skipped_files.append((file_path, str(e)))
            print(f"Error processing {file_path}: {e}")
    
    # Save the Excel file
    try:
        writer.close()  # Changed from writer.save() to writer.close()
        print(f"Successfully saved Excel file to: {output_excel}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return imported_count, len(file_paths), None
    
    # Print report
    print(f"\nSuccessfully imported {imported_count} files to {output_excel}")
    
    if skipped_files:
        print(f"\nSkipped {len(skipped_files)} files:")
        for file_path, error in skipped_files:
            print(f"- {os.path.basename(file_path)}: {error}")
    
    return imported_count, len(file_paths), output_excel


def get_files_from_directory(directory_path):
    """Find all CSV and TSV files in a directory."""
    csv_files = glob.glob(os.path.join(directory_path, '*.csv'))
    tsv_files = glob.glob(os.path.join(directory_path, '*.tsv'))
    return sorted(csv_files + tsv_files)


def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Convert multiple CSV/TSV files to Excel sheets')
    
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument('-i', '--input', nargs='+', 
                        help='CSV/TSV files to import or a directory containing CSV/TSV files')
    input_group.add_argument('-I', '--input-file', 
                        help='File containing a list of input files (one per line)')
    
    parser.add_argument('-o', '--output', required=True, 
                        help='Path to save the Excel file')
    parser.add_argument('--horizontal-headers', action='store_true',
                        help='Keep headers horizontal (default is vertical)')
    parser.add_argument('--heatmap', action='store_true',
                        help='Apply heatmap coloring to numeric data cells')
    parser.add_argument('--skip-heatmap-cols', nargs='+', default=[],
                        help='Column names to exclude from heatmap coloring')
    
    args = parser.parse_args()
    
    # Collect input files
    file_paths = []
    
    # If we have an input file list
    if args.input_file:
        try:
            with open(args.input_file, 'r') as f:
                for line in f:
                    path = line.strip()
                    if path and not path.startswith('#'):  # Skip empty lines and comments
                        file_paths.append(path)
        except Exception as e:
            print(f"Error reading input file list: {e}")
            sys.exit(1)
    else:
        # Process input arguments (files or directories)
        for path in args.input:
            if os.path.isdir(path):
                # If directory, add all CSV/TSV files
                dir_files = get_files_from_directory(path)
                if not dir_files:
                    print(f"No CSV/TSV files found in directory: {path}")
                file_paths.extend(dir_files)
            elif os.path.isfile(path):
                # If file, add it directly
                file_paths.append(path)
            else:
                print(f"Warning: {path} not found, skipping.")
    
    # Check if we have files to process
    if not file_paths:
        print("No valid input files found. Exiting.")
        sys.exit(1)
    
    print(f"Found {len(file_paths)} files to process")
    
    # Import files
    imported_count, total_count, output_path = import_files_to_excel(
        file_paths, 
        args.output,
        horizontal_headers=args.horizontal_headers,
        apply_heatmap_coloring=args.heatmap,
        skip_heatmap_cols=args.skip_heatmap_cols
    )
    
    # Show final result
    if imported_count > 0:
        print(f"\nSuccess: Imported {imported_count}/{total_count} files to {output_path}")
        return 0
    else:
        print("\nFailed: No files were successfully imported.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
